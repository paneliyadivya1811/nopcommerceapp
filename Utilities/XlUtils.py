import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheetName]
    return(Sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheetName]
    return(Sheet.max_column)

def Readdata(file, sheetName, rowNo, columnNo):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheetName]
    return Sheet.cell(rowNo, columnNo).value

def Writedata(file, sheetName, rowNo, columnNo, data):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheetName]
    Sheet.cell(rowNo, columnNo).value = data
    workbook.save(file)

def fillGreenColor(file, sheetName, rowNo, columnNo):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheetName]
    Greenfill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    Sheet.cell(rowNo, columnNo).fill = Greenfill
    workbook.save(file)

def fillRedColor(file, sheetName, rowNo, columnNo):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[sheetName]
    Redfill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    Sheet.cell(rowNo, columnNo).fill = Redfill
    workbook.save(file)
